from flask import Flask, render_template, jsonify, send_file, request
import psutil
import platform
import speedtest
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

app = Flask(__name__)

def obtener_caracteristicas_equipo():
    """Obtiene las características del equipo"""
    try:
        # Información del sistema
        sistema = platform.system()
        version = platform.version()
        arquitectura = platform.architecture()[0]
        
        # Información del procesador
        procesador = platform.processor()
        nucleos_fisicos = psutil.cpu_count(logical=False)
        nucleos_logicos = psutil.cpu_count(logical=True)
        
        # Información de memoria
        memoria = psutil.virtual_memory()
        ram_gb = round(memoria.total / (1024**3), 1)
        
        # Información del disco
        disco = psutil.disk_usage('/')
        disco_gb = round(disco.total / (1024**3), 2)
        
        # Evaluar estado del equipo (simplificado)
        estado_equipo = "APROBADO" if ram_gb >= 4 and nucleos_fisicos >= 2 else "RECHAZADO"
        
        return {
            'sistema_operativo': f"{sistema} {version}",
            'arquitectura': arquitectura,
            'procesador': procesador,
            'nucleos_fisicos': nucleos_fisicos,
            'nucleos_logicos': nucleos_logicos,
            'ram': f"{ram_gb} GB",
            'disco': f"{disco_gb} GB",
            'estado_equipo': estado_equipo
        }
    except Exception as e:
        return {
            'error': f"Error al obtener características del equipo: {str(e)}"
        }

def medir_velocidad_internet():
    """Mide la velocidad de internet"""
    try:
        # Intentar usar speedtest-cli
        try:
            st = speedtest.Speedtest()
            st.get_best_server()
            
            # Medir velocidad de descarga
            velocidad_descarga = st.download() / 1_000_000  # Convertir a Mbps
            
            # Medir velocidad de carga
            velocidad_carga = st.upload() / 1_000_000  # Convertir a Mbps
            
            # Obtener latencia
            latencia = st.results.ping
            
            # Evaluar estado de internet
            estado_internet = "APROBADO" if velocidad_descarga >= 10 and velocidad_carga >= 5 else "RECHAZADO"
            
            return {
                'velocidad_descarga': round(velocidad_descarga, 2),
                'velocidad_carga': round(velocidad_carga, 2),
                'latencia': round(latencia, 2),
                'estado_internet': estado_internet
            }
        except Exception as speedtest_error:
            print(f"Speedtest error: {speedtest_error}")
            # Fallback: medición básica usando requests
            return medir_velocidad_basica()
            
    except Exception as e:
        print(f"Error general en medición de internet: {e}")
        return medir_velocidad_basica()

def medir_velocidad_basica():
    """Medición básica de velocidad usando requests"""
    try:
        import requests
        import time
        
        # Medir latencia
        start_time = time.time()
        response = requests.get('https://www.google.com', timeout=10)
        end_time = time.time()
        latencia = round((end_time - start_time) * 1000, 2)  # ms
        
        # Medir velocidad de descarga
        download_start = time.time()
        download_response = requests.get('https://httpbin.org/bytes/1048576', timeout=30)  # 1MB
        download_end = time.time()
        download_time = download_end - download_start
        download_size = len(download_response.content) * 8  # bits
        velocidad_descarga = (download_size / download_time) / 1_000_000  # Mbps
        
        # Medir velocidad de carga
        upload_start = time.time()
        test_data = 'x' * 102400  # 100KB
        upload_response = requests.post('https://httpbin.org/post', data=test_data, timeout=30)
        upload_end = time.time()
        upload_time = upload_end - upload_start
        upload_size = len(test_data) * 8  # bits
        velocidad_carga = (upload_size / upload_time) / 1_000_000  # Mbps
        
        # Evaluar estado de internet
        estado_internet = "APROBADO" if velocidad_descarga >= 10 and velocidad_carga >= 5 else "RECHAZADO"
        
        return {
            'velocidad_descarga': round(velocidad_descarga, 2),
            'velocidad_carga': round(velocidad_carga, 2),
            'latencia': latencia,
            'estado_internet': estado_internet
        }
    except Exception as e:
        print(f"Error en medición básica: {e}")
        return {
            'velocidad_descarga': 0,
            'velocidad_carga': 0,
            'latencia': 0,
            'estado_internet': 'ERROR'
        }

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/evaluar', methods=['GET', 'POST'])
def evaluar():
    """Endpoint para confirmar que la evaluación se realiza del lado del cliente"""
    return jsonify({
        'message': 'La evaluación se realiza del lado del cliente',
        'status': 'success'
    })

def obtener_caracteristicas_equipo_servidor(info_cliente):
    """Obtiene características del equipo usando información del servidor y cliente"""
    try:
        # Información del sistema del servidor (más precisa)
        sistema = platform.system()
        version = platform.version()
        arquitectura = platform.architecture()[0]
        
        # Información del procesador del servidor
        procesador = platform.processor()
        nucleos_fisicos = psutil.cpu_count(logical=False)
        nucleos_logicos = psutil.cpu_count(logical=True)
        
        # Información de memoria del servidor
        memoria = psutil.virtual_memory()
        ram_gb = round(memoria.total / (1024**3), 1)
        
        # Información del disco del servidor
        disco = psutil.disk_usage('/')
        disco_gb = round(disco.total / (1024**3), 2)
        
        # Usar información del cliente para mejorar la detección
        user_agent = info_cliente.get('userAgent', '')
        
        # Mejorar detección del sistema operativo
        if 'Windows NT 10.0' in user_agent:
            sistema_operativo = 'Windows 11'
        elif 'Windows NT 6.3' in user_agent:
            sistema_operativo = 'Windows 8.1'
        elif 'Windows NT 6.2' in user_agent:
            sistema_operativo = 'Windows 8'
        elif 'Windows NT 6.1' in user_agent:
            sistema_operativo = 'Windows 7'
        elif 'Mac OS X' in user_agent:
            sistema_operativo = 'macOS'
        elif 'Linux' in user_agent:
            sistema_operativo = 'Linux'
        else:
            sistema_operativo = f"{sistema} {version}"
        
        # Obtener información real del procesador
        procesador_info = "No disponible"
        try:
            import subprocess
            if sistema == 'Windows':
                # Obtener información completa del CPU en Windows
                result = subprocess.run(['wmic', 'cpu', 'get', 'name,numberofcores,numberoflogicalprocessors,maxclockspeed', '/format:csv'], 
                                      capture_output=True, text=True, shell=True)
                if result.returncode == 0:
                    lines = result.stdout.strip().split('\n')
                    if len(lines) > 1:
                        # Parsear la información del CPU
                        cpu_info = lines[1].split(',')
                        if len(cpu_info) >= 4:
                            cpu_name = cpu_info[1].strip()
                            cores = cpu_info[2].strip()
                            logical_cores = cpu_info[3].strip()
                            max_speed = cpu_info[4].strip()
                            
                            # Formatear según el formato solicitado
                            if 'Intel' in cpu_name and 'Core' in cpu_name:
                                # Extraer generación y modelo
                                if '12th' in cpu_name:
                                    generacion = '12th Gen'
                                elif '11th' in cpu_name:
                                    generacion = '11th Gen'
                                elif '10th' in cpu_name:
                                    generacion = '10th Gen'
                                else:
                                    generacion = 'Intel'
                                
                                # Extraer modelo específico
                                if 'i7' in cpu_name:
                                    if '1260P' in cpu_name:
                                        modelo = 'i7-1260P'
                                    elif '1270P' in cpu_name:
                                        modelo = 'i7-1270P'
                                    else:
                                        modelo = 'i7'
                                elif 'i5' in cpu_name:
                                    modelo = 'i5'
                                elif 'i3' in cpu_name:
                                    modelo = 'i3'
                                elif 'i9' in cpu_name:
                                    modelo = 'i9'
                                else:
                                    modelo = 'Processor'
                                
                                # Calcular frecuencia aproximada
                                freq_mhz = int(max_speed) if max_speed.isdigit() else 2100
                                freq_ghz = round(freq_mhz / 1000, 1)
                                
                                procesador_info = f"{generacion} Intel(R) Core(TM) {modelo} ({logical_cores} CPUs), ~{freq_ghz}GHz"
                            else:
                                procesador_info = f"{cpu_name} ({logical_cores} CPUs), ~{round(int(max_speed)/1000, 1) if max_speed.isdigit() else 2.0}GHz"
                        else:
                            procesador_info = f"{nucleos_logicos} cores"
                    else:
                        procesador_info = f"{nucleos_logicos} cores"
                else:
                    procesador_info = f"{nucleos_logicos} cores"
            else:
                # Para Linux/Mac
                try:
                    result = subprocess.run(['lscpu'], capture_output=True, text=True)
                    if result.returncode == 0:
                        lines = result.stdout.split('\n')
                        cpu_model = 'Unknown'
                        cpu_cores = nucleos_logicos
                        cpu_freq = '2.0'
                        
                        for line in lines:
                            if 'Model name:' in line:
                                cpu_model = line.split(':')[1].strip()
                            elif 'CPU(s):' in line:
                                cpu_cores = line.split(':')[1].strip()
                            elif 'CPU max MHz:' in line:
                                freq_mhz = line.split(':')[1].strip()
                                if freq_mhz.isdigit():
                                    cpu_freq = str(round(int(freq_mhz) / 1000, 1))
                        
                        procesador_info = f"{cpu_model} ({cpu_cores} CPUs), ~{cpu_freq}GHz"
                    else:
                        procesador_info = f"{nucleos_logicos} cores"
                except:
                    procesador_info = f"{nucleos_logicos} cores"
        except Exception as e:
            print(f"Error obteniendo información del procesador: {e}")
            procesador_info = f"{nucleos_logicos} cores"
        
        # Evaluar estado del equipo
        estado_equipo = "APROBADO" if ram_gb >= 4 and nucleos_fisicos >= 2 else "RECHAZADO"
        
        return {
            'sistema_operativo': sistema_operativo,
            'arquitectura': arquitectura,
            'procesador': procesador_info,
            'nucleos_fisicos': nucleos_fisicos,
            'nucleos_logicos': nucleos_logicos,
            'ram': f"{ram_gb} GB",
            'disco': f"{disco_gb} GB",
            'estado_equipo': estado_equipo
        }
    except Exception as e:
        return {
            'sistema_operativo': 'Error al detectar',
            'arquitectura': 'Error al detectar',
            'procesador': 'Error al detectar',
            'nucleos_fisicos': 'Error al detectar',
            'nucleos_logicos': 'Error al detectar',
            'ram': 'Error al detectar',
            'disco': 'Error al detectar',
            'estado_equipo': 'ERROR'
        }

@app.route('/exportar-excel', methods=['POST'])
def exportar_excel():
    """Exporta los resultados a un archivo Excel con datos del cliente"""
    try:
        # Obtener datos del cliente
        data = request.get_json()
        usuario = data.get('usuario', 'Usuario')
        fecha = data.get('fecha', datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
        caracteristicas = data.get('equipo', {})
        velocidad = data.get('internet', {})
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultado Evaluación"
        
        # Estilos
        titulo_font = Font(bold=True, size=14, color="FFFFFF")
        subtitulo_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # Título principal
        ws['A1'] = "RESULTADO DE EVALUACIÓN"
        ws['A1'].font = titulo_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:C1')
        
        # Usuario y fecha
        ws['A2'] = f"Usuario: {usuario}"
        ws.merge_cells('A2:C2')
        ws['A3'] = f"Fecha: {fecha}"
        ws.merge_cells('A3:C3')
        
        # Evaluación del Equipo
        ws['A5'] = "EVALUACIÓN DEL EQUIPO"
        ws['A5'].font = subtitulo_font
        ws['A5'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A5'].font = Font(bold=True, color="FFFFFF")
        ws.merge_cells('A5:C5')
        
        # Datos del equipo
        datos_equipo = [
            ["Sistema Operativo", caracteristicas.get('sistema_operativo', 'N/A')],
            ["Arquitectura", caracteristicas.get('arquitectura', 'N/A')],
            ["Procesador", caracteristicas.get('procesador', 'N/A')],
            ["RAM", caracteristicas.get('ram', 'N/A')],
            ["Disco", caracteristicas.get('disco', 'N/A')],
            ["Estado del equipo", caracteristicas.get('estado_equipo', 'N/A')]
        ]
        
        for i, (campo, valor) in enumerate(datos_equipo, start=6):
            ws[f'A{i}'] = campo
            ws[f'B{i}'] = valor
            if campo == "Estado del equipo":
                if valor == "APROBADO":
                    ws[f'B{i}'].fill = green_fill
                    ws[f'B{i}'].font = Font(bold=True, color="FFFFFF")
        
        # Evaluación de Internet
        ws['A15'] = "EVALUACIÓN DE INTERNET"
        ws['A15'].font = subtitulo_font
        ws['A15'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws['A15'].font = Font(bold=True, color="FFFFFF")
        ws.merge_cells('A15:C15')
        
        # Datos de internet
        datos_internet = [
            ["Velocidad de descarga", f"{velocidad.get('velocidad_descarga', 'N/A')} Mbps"],
            ["Velocidad de carga", f"{velocidad.get('velocidad_carga', 'N/A')} Mbps"],
            ["Latencia", f"{velocidad.get('latencia', 'N/A')} ms"],
            ["Estado de internet", velocidad.get('estado_internet', 'N/A')]
        ]
        
        for i, (campo, valor) in enumerate(datos_internet, start=16):
            ws[f'A{i}'] = campo
            ws[f'B{i}'] = valor
            if campo == "Estado de internet":
                if valor == "APROBADO":
                    ws[f'B{i}'].fill = green_fill
                    ws[f'B{i}'].font = Font(bold=True, color="FFFFFF")
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        
        # Guardar en memoria
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'evaluacion_{usuario}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': f'Error al exportar Excel: {str(e)}'}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)

